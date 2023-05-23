import openpyxl
import requests
import telebot
import time as t
import datetime
import sqlite3
import threading
import schedule
import json
from telebot import types
from datetime import date, time

bot_token = '6224714194:AAERfVMEJn87MunWtYcH8MAtAOkRVelPynE'
bot = telebot.TeleBot(bot_token)

excel_file_path = 'kmb212(1).xlsx'
excel_file_path2 = 'D:\kmb211.xlsx'


conn = sqlite3.connect('id.db', check_same_thread=False)
cursor = conn.cursor()
cursor.execute('''CREATE TABLE IF NOT EXISTS kmb211
                (id INTEGER PRIMARY KEY)''')
cursor.execute('''CREATE TABLE IF NOT EXISTS kmb212
                (id INTEGER PRIMARY KEY)''')
cursor.execute('''CREATE TABLE IF NOT EXISTS kmb221
                (id INTEGER PRIMARY KEY)''')
cursor.execute('''CREATE TABLE IF NOT EXISTS kmb222
                (id INTEGER PRIMARY KEY)''')
cursor.execute('''CREATE TABLE IF NOT EXISTS its211
                (id INTEGER PRIMARY KEY)''')
cursor.execute('''CREATE TABLE IF NOT EXISTS its212
                (id INTEGER PRIMARY KEY)''')




@bot.message_handler(commands=['start'])
def start_command(message):

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

    item1 = types.KeyboardButton("КМБ-21-1")
    item2 = types.KeyboardButton("КМБ-22-1")
    item3 = types.KeyboardButton("ИТС-21-1")
    markup.add(item1, item2, item3)

    bot.send_message(message.chat.id, "Привет! Нажми на кнопку, чтобы выбрать свою академическую группу", reply_markup=markup)


@bot.message_handler(content_types='text')
def message_reply(message):

    markup2 = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup3 = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup4 = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup5 = types.ReplyKeyboardMarkup(resize_keyboard=True)

    item4 = types.KeyboardButton(text="КМБ-21-1(1)")
    item5 = types.KeyboardButton(text="КМБ-21-1(2)")
    markup2.add(item4, item5)

    item6 = types.KeyboardButton("КМБ-22-1(1)")
    item7 = types.KeyboardButton("КМБ-22-1(2)")
    markup3.add(item6, item7)

    item8 = types.KeyboardButton("ИТС-21-1(1)")
    item9 = types.KeyboardButton("ИТС-21-1(2)")
    markup4.add(item8, item9)

    item10 = types.KeyboardButton("Сегодня")
    item11 = types.KeyboardButton("Завтра")
    markup5.add(item10, item11)

    text = message.text

    r=""
    r2 = ""
    r3 = ""
    r4 = ""
    r5 = ""
    r6 = ""



    if message.text == "КМБ-21-1":
        bot.send_message(message.chat.id, "Выберите подгруппу", reply_markup=markup2)
    if message.text == "КМБ-21-1(1)":
        print(text)
        bot.send_message(message.chat.id, "Теперь вам будет приходить ваше расписание",
                         reply_markup=markup5)
        # global chat_id
        chat_id = message.chat.id
        # проверяем, есть ли уже ID чата в базе данных
        cursor.execute('SELECT * FROM kmb211 WHERE id = ?', (chat_id,))
        result = cursor.fetchone()
        r=result
        if result is None:
            cursor.execute('INSERT INTO kmb211 VALUES (?)', (chat_id,))
            conn.commit()

        #     bot.send_message(message.chat.id, 'ID чата сохранен в базе данных')
        # else:
        #     bot.send_message(message.chat.id, 'ID чата уже существует в базе данных')
    if text == "Сегодня" and r != None:
        day = date.today()
        week_number = day.weekday()
        work = openpyxl.load_workbook("kmb212(1).xlsx")
        kmb211 = work.worksheets[1]
        if week_number == 0:
            bot.send_message(message.chat.id, kmb211['A1'].value)
        if week_number == 1:
            bot.send_message(message.chat.id, kmb211['A2'].value)
        if week_number == 2:
            bot.send_message(message.chat.id, kmb211['A3'].value)
        if week_number == 3:
            bot.send_message(message.chat.id, kmb211['A4'].value)
        if week_number == 4:
            bot.send_message(message.chat.id, kmb211['A5'].value)
        if week_number == 5:
            bot.send_message(message.chat.id, kmb211['A6'].value)
    if text == "Завтра" and r != None:
        day = date.today()
        week_number = day.weekday()
        work = openpyxl.load_workbook("kmb212(1).xlsx")
        kmb211 = work.worksheets[1]
        if week_number == 0:
            bot.send_message(message.chat.id, kmb211['A2'].value)
        if week_number == 1:
            bot.send_message(message.chat.id, kmb211['A3'].value)
        if week_number == 2:
            bot.send_message(message.chat.id, kmb211['A4'].value)
        if week_number == 3:
            bot.send_message(message.chat.id, kmb211['A5'].value)
        if week_number == 4:
            bot.send_message(message.chat.id, kmb211['A6'].value)
        if week_number == 5:
            bot.send_message(message.chat.id, kmb211['A7'].value)




        # if weekday_number == 4:
        #     if current_time == "17:41":
        #         wb = openpyxl.load_workbook(excel_file_path)
        #         ws = wb.active
        #         data = ws['A1'].value
        #         bot.send_message(message.chat.id, data)
        #     if current_time == "17:42":
        #         wb = openpyxl.load_workbook(excel_file_path)
        #         ws = wb.active
        #         data = ws['A1'].value
        #         bot.send_message(message.chat.id, data)
        #     if current_time == "17:43":
        #         wb = openpyxl.load_workbook(excel_file_path)
        #         ws = wb.active
        #         data = ws['A1'].value
        #         bot.send_message(message.chat.id, data)


    if message.text == "КМБ-21-1(2)":
        bot.send_message(message.chat.id, "Теперь вам будет приходить ваше расписание", reply_markup=markup5)

        chat_id = message.chat.id
        # проверяем, есть ли уже ID чата в базе данных
        cursor.execute('SELECT * FROM kmb212 WHERE id = ?', (chat_id,))
        result2 = cursor.fetchone()
        r2 = result2
        if result2 is None:
            cursor.execute('INSERT INTO kmb212 VALUES (?)', (chat_id,))
            conn.commit()

    if text == "Сегодня" and r2 != None:
        day = date.today()
        week_number = day.weekday()
        work = openpyxl.load_workbook("kmb212(1).xlsx")
        kmb212 = work.worksheets[0]
        if week_number == 0:
            bot.send_message(message.chat.id, kmb212['A1'].value)
        if week_number == 1:
            bot.send_message(message.chat.id, kmb212['A2'].value)
        if week_number == 2:
            bot.send_message(message.chat.id, kmb212['A3'].value)
        if week_number == 3:
            bot.send_message(message.chat.id, kmb212['A4'].value)
        if week_number == 4:
            bot.send_message(message.chat.id, kmb212['A5'].value)
        if week_number == 5:
            bot.send_message(message.chat.id, kmb212['A6'].value)
    if text == "Завтра" and r2 != None:
        day = date.today()
        week_number = day.weekday()
        work = openpyxl.load_workbook("kmb212(1).xlsx")
        kmb212 = work.worksheets[0]
        if week_number == 0:
            bot.send_message(message.chat.id, kmb212['A2'].value)
        if week_number == 1:
            bot.send_message(message.chat.id, kmb212['A3'].value)
        if week_number == 2:
            bot.send_message(message.chat.id, kmb212['A4'].value)
        if week_number == 3:
            bot.send_message(message.chat.id, kmb212['A5'].value)
        if week_number == 4:
            bot.send_message(message.chat.id, kmb212['A6'].value)
        if week_number == 5:
            bot.send_message(message.chat.id, kmb212['A7'].value)

        # if weekday_number == 6:
        #     if time == "18:27":
        #         wb = openpyxl.load_workbook(excel_file_path)
        #         ws = wb.active
        #         data = ws['A1'].value
        #         bot.send_message(message.chat.id, data)


    if message.text == "КМБ-22-1":
        bot.send_message(message.chat.id, "Выберите подгруппу", reply_markup=markup3)


    if message.text == "ИТС-21-1":
        bot.send_message(message.chat.id, "Выберите подгруппу", reply_markup=markup4)
    if message.text == "ИТС-21-1(1)":
        bot.send_message(message.chat.id, "Теперь вам будет приходить ваше расписание", reply_markup=types.ReplyKeyboardRemove())
        chat_id = message.chat.id
        # проверяем, есть ли уже ID чата в базе данных
        cursor.execute('SELECT * FROM its211 WHERE id = ?', (chat_id,))
        result5 = cursor.fetchone()
        r5 = result5
        if result5 is None:
            cursor.execute('INSERT INTO its211 VALUES (?)', (chat_id,))
            conn.commit()
    if message.text == "ИТС-21-1(2)":
        bot.send_message(message.chat.id, "Теперь вам будет приходить ваше расписание", reply_markup=types.ReplyKeyboardRemove())
        chat_id = message.chat.id
        # проверяем, есть ли уже ID чата в базе данных
        cursor.execute('SELECT * FROM its212 WHERE id = ?', (chat_id,))

        result6 = cursor.fetchone()
        r6 = result6
        if result6 is None:
            cursor.execute('INSERT INTO its212 VALUES (?)', (chat_id,))
            conn.commit()


def worker():
    while 1:
        today = date.today()

        weekday_number = today.weekday()
        current_time = datetime.datetime.now().strftime('%H:%M:%S')
        # print(current_time)
        # print(weekday_number)
        cou = 1
        wb = openpyxl.load_workbook("kmb212(1).xlsx")
        kmb212 = wb.worksheets[0]
        kmb211 = wb.worksheets[1]
        kmb221 = wb.worksheets[2]
        kmb222 = wb.worksheets[3]
        its211 = wb.worksheets[4]
        its212 = wb.worksheets[5]
        para1 = kmb212['B1'].value
        para2 = kmb212['C1'].value
        para3 = kmb212['D1'].value
        para4 = kmb212['E1'].value
        para5 = kmb212['F1'].value
        para6 = kmb212['G1'].value
        para7 = kmb212['H1'].value


        cursor22 = conn.execute("SELECT id FROM kmb212")
        user_id2 = [row[0] for row in cursor22.fetchall()]
        for id in user_id2:
            if cou % 2 != 0:
                if weekday_number == 0:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb212['A1'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == str(para5):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 1:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb212['A2'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == "18:49:20":
                        bot.send_message(id, 'Привет, напиши мне в вк если это сообщение дойдёт')
                        t.sleep(3)
                if weekday_number == 2:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb212['A3'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == str(para5):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 3:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb212['A4'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == str(
                            para5) or current_time == str(para2):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 4:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb212['A5'].value)
                    if current_time == str(para3) or current_time == str(para2) or current_time == str(para4):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 5:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb212['A6'].value)
                    if current_time == str(para4) or current_time == str(para5):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)

                if weekday_number == 6:
                    cou = +1
            else:
                if weekday_number == 0:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb212['I1'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == str(para5):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 1:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb212['I2'].value)
                    if current_time == str(para3) or current_time == str(para4):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 2:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb212['I3'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == str(para5):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 3:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb212['I4'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == str(
                            para5) or current_time == str(para2):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 4:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb212['I5'].value)
                    if current_time == str(para3) or current_time == str(para2) or current_time == str(para4):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)


                if weekday_number == 6:
                    cou = +1


        cursor21 = conn.execute("SELECT id FROM kmb211")
        user_id = [row[0] for row in cursor21.fetchall()]
        for id in user_id:
            if cou%2 != 0:
                if weekday_number == 0:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb211['A1'].value)
                    if current_time == str(para4) or current_time == str(para5) or current_time == str(para6) or current_time == "13:18:40":
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 1:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb211['A2'].value)
                    if current_time == str(para3) or current_time ==str(para4) or current_time ==str(para5) or current_time == "18:49:20":
                        bot.send_message(id, 'Привет, напиши мне в вк если это сообщение дойдёт')
                        t.sleep(3)
                if weekday_number == 2:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb211['A3'].value)
                    if current_time == str(para2) or current_time == str(para3) or current_time == str(para4):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 3:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb211['A4'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == str(para5) or current_time == str(para2):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 4:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb211['A5'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == str(para5):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)


                if weekday_number == 6:
                    cou=+1
            else:
                if weekday_number == 0:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb211['B1'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == str(para5):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 1:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb211['B2'].value)
                    if current_time == str(para3) or current_time == str(para4):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 2:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb211['B3'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == str(para5):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 3:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb211['B4'].value)
                    if current_time == str(para3) or current_time == str(para4) or current_time == str(
                            para5) or current_time == str(para2):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 4:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb211['B5'].value)
                    if current_time == str(para3) or current_time == str(para2):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)
                if weekday_number == 5:
                    if current_time == kmb212['B2'].value:
                        bot.send_message(id, kmb211['B6'].value)
                    if current_time == str(para4) or current_time == str(para5):
                        bot.send_message(id, 'Пара начнётся через 5 минут')
                        t.sleep(3)

                if weekday_number == 6:
                    cou = +1



thread = threading.Thread(target=worker)
thread.start()
bot.infinity_polling()
